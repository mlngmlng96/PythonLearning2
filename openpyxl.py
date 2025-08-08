import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename,sheetname):
    wb = xl.load_workbook(filename)
    sheet = wb[sheetname]

    for row in range(2, sheet.max_row + 1): # add 1 to include the last row
        price = sheet.cell(row,3)
        corrected_price = price.value * 0.9
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
              min_row = 2,
              max_row = sheet.max_row,
              min_col = 4,
              max_col = 4
              )

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')

    wb.save(filename)


process_workbook('transactions.xlsx','Sheet1')


